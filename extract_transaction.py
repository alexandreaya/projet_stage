from web3 import Web3
import openpyxl


infura_url = 'https://goerli.infura.io/v3/91e6692c109c47c395f5d115a5a01395'
web3 = Web3(Web3.HTTPProvider(infura_url))

# Liste des transactions (hash)
transaction_hashes = [
'0xd7ce5177dc32c0f5fa7a3c5d5c752a3592ed9906c8fd1291190ec9e1e133890d'



]

# Création du fichier Excel

excel_file_path = 'transaction_data.xlsx'
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'Block Number'
sheet['B1'] = 'Transaction Hash'
sheet['C1'] = 'From Address'
sheet['D1'] = 'To Address'
sheet['E1'] = 'Gas'
sheet['F1'] = 'Gas Price'
sheet['G1'] = 'Max Fee Per Gas'
sheet['H1'] = 'Max Priority Fee Per Gas'
sheet['I1'] = 'Value'

# Parcourir les transactions et extraire les informations
for i, transaction_hash in enumerate(transaction_hashes):
    transaction = web3.eth.get_transaction(transaction_hash)

    # Vérifier si la transaction est valide
    if transaction is not None:
        block_number = transaction.blockNumber
        transaction_hash = transaction.hash.hex()
        from_address = transaction['from']
        to_address = transaction['to']
        gas = transaction['gas']
        gas_price = transaction['gasPrice']
        max_fee_per_gas = transaction['maxFeePerGas']
        max_priority_fee_per_gas = transaction['maxPriorityFeePerGas']
        value = web3.from_wei(transaction['value'], 'ether')

        # Ajout des données de la transaction
        row = i + 2
        sheet.cell(row=row, column=1, value=block_number)
        sheet.cell(row=row, column=2, value=transaction_hash)
        sheet.cell(row=row, column=3, value=from_address)
        sheet.cell(row=row, column=4, value=to_address)
        sheet.cell(row=row, column=5, value=gas)
        sheet.cell(row=row, column=6, value=gas_price)
        sheet.cell(row=row, column=7, value=max_fee_per_gas)
        sheet.cell(row=row, column=8, value=max_priority_fee_per_gas)
        sheet.cell(row=row, column=9, value=value)

    else:
        print('hash invalide:', transaction_hash)

# Enregistrement du fichier Excel
workbook.save(excel_file_path)

print('La transaction a bien été extraite', excel_file_path)
